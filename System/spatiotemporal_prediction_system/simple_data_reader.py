"""
简单数据读取模块 - 不依赖pandas，使用openpyxl直接读取Excel
"""
import os
import json
from datetime import datetime

class SimpleDataReader:
    """简单数据读取器 - 用于读取原始数据和预测数据"""
    
    def __init__(self, base_dir="."):
        self.base_dir = base_dir
        self.raw_data_path = os.path.join(base_dir, "时序数据", "原始数据.xlsx")
        self.prediction_data_dir = os.path.join(base_dir, "时序数据")
        
        # 模型名称到文件名的映射
        self.model_files = {
            "CNN-LSTM-Attention": "CNN-LSTM-Attention-预测数据.xlsx",
            "GRU": "GRU-预测数据.xlsx",
            "KAN": "KAN-预测数据.xlsx",
            "KNN": "KNN-预测数据.xlsx",
            "LSTM": "LSTM-预测数据.xlsx",
            "PatchFormer": "PatchFormer-预测数据.xlsx",
            "PatchTST": "PatchTST-预测数据.xlsx",
            "SVR": "SVR-预测数据.xlsx",
            "TCN": "TCN-预测数据.xlsx",
            "TimesNet": "TimesNet-预测数据.xlsx",
            "TSPeakNet": "TSPeakNet-预测模型.xlsx"
        }
    
    def read_raw_data(self, limit=1000):
        """读取原始数据 - 使用openpyxl"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self.raw_data_path, data_only=True)
            ws = wb.active
            
            # 读取表头
            headers = []
            for cell in ws[1]:
                headers.append(cell.value if cell.value else "")
            
            # 读取数据行
            data = []
            for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, limit+1)):
                row_data = {}
                for idx, cell in enumerate(row):
                    if idx < len(headers):
                        value = cell.value
                        # 处理日期格式
                        if isinstance(value, datetime):
                            value = value.strftime('%Y-%m-%d')
                        row_data[headers[idx]] = value
                data.append(row_data)
            
            return {
                "status": "success",
                "headers": headers,
                "data": data,
                "total_rows": ws.max_row - 1
            }
        except Exception as e:
            return {
                "status": "error",
                "message": str(e),
                "headers": [],
                "data": [],
                "total_rows": 0
            }
    
    def get_yearly_statistics(self):
        """获取年度统计数据"""
        result = self.read_raw_data(limit=10000)
        if result["status"] != "success":
            return []
        
        # 统计年度数据
        yearly_data = {}
        for row in result["data"]:
            date_str = row.get("开具时间", "") or row.get("日期", "") or row.get("Date", "")
            if not date_str:
                continue
            try:
                if isinstance(date_str, str):
                    year = int(date_str[:4]) if len(date_str) >= 4 else None
                else:
                    continue
                if year:
                    if year not in yearly_data:
                        yearly_data[year] = {"area": 0, "count": 0, "total_value": 0}
                    yearly_data[year]["count"] += 1
                    
                    # 尝试计算各区域的总值
                    total_value = 0
                    for col in result["headers"]:
                        if col.startswith("Node_") or col in ["AT", "MaxT", "MinT", "AWS", "MSWS", "ADP", "Precip"]:
                            val = row.get(col)
                            if val is not None:
                                try:
                                    total_value += float(val)
                                except:
                                    pass
                    yearly_data[year]["total_value"] += total_value
            except:
                continue
        
        # 转换为列表
        yearly_list = []
        for year in sorted(yearly_data.keys()):
            yearly_list.append({
                "year": year,
                "count": yearly_data[year]["count"],
                "average": yearly_data[year]["total_value"] / yearly_data[year]["count"] if yearly_data[year]["count"] > 0 else 0
            })
        
        return yearly_list
    
    def get_monthly_statistics(self):
        """获取月度统计数据"""
        result = self.read_raw_data(limit=10000)
        if result["status"] != "success":
            return []
        
        # 统计月度数据
        monthly_data = {}
        for row in result["data"]:
            date_str = row.get("开具时间", "") or row.get("日期", "") or row.get("Date", "")
            if not date_str:
                continue
            try:
                if isinstance(date_str, str) and len(date_str) >= 7:
                    year_month = date_str[:7]
                    month_num = int(date_str[5:7])
                    if year_month not in monthly_data:
                        monthly_data[year_month] = {
                            "year": int(date_str[:4]),
                            "month": month_num,
                            "count": 0,
                            "total_value": 0
                        }
                    monthly_data[year_month]["count"] += 1
                    
                    # 计算总数值
                    total_value = 0
                    for col in result["headers"]:
                        if col.startswith("Node_") or col in ["AT", "MaxT", "MinT", "AWS", "MSWS", "ADP", "Precip"]:
                            val = row.get(col)
                            if val is not None:
                                try:
                                    total_value += float(val)
                                except:
                                    pass
                    monthly_data[year_month]["total_value"] += total_value
            except:
                continue
        
        # 转换为列表并排序
        monthly_list = []
        for key in sorted(monthly_data.keys()):
            avg = monthly_data[key]["total_value"] / monthly_data[key]["count"] if monthly_data[key]["count"] > 0 else 0
            monthly_list.append({
                "year": monthly_data[key]["year"],
                "month": monthly_data[key]["month"],
                "count": monthly_data[key]["count"],
                "average": avg
            })
        
        return monthly_list
    
    def get_regional_statistics(self):
        """获取区域统计数据"""
        result = self.read_raw_data(limit=10000)
        if result["status"] != "success":
            return []
        
        # 统计区域数据 - 按Node节点统计
        regional_data = {}
        
        # 先找所有区域节点
        node_columns = [col for col in result["headers"] if col.startswith("Node_")]
        
        for row in result["data"]:
            for node_col in node_columns:
                node_name = node_col.replace("Node_", "")
                if node_name not in regional_data:
                    regional_data[node_name] = {
                        "county": node_name,
                        "count": 0,
                        "total_value": 0,
                        "values": []
                    }
                
                val = row.get(node_col)
                if val is not None:
                    try:
                        float_val = float(val)
                        regional_data[node_name]["total_value"] += float_val
                        regional_data[node_name]["values"].append(float_val)
                        regional_data[node_name]["count"] += 1
                    except:
                        pass
        
        # 转换为列表
        regional_list = []
        for node_name in regional_data.keys():
            avg = regional_data[node_name]["total_value"] / regional_data[node_name]["count"] if regional_data[node_name]["count"] > 0 else 0
            max_val = max(regional_data[node_name]["values"]) if regional_data[node_name]["values"] else 0
            min_val = min(regional_data[node_name]["values"]) if regional_data[node_name]["values"] else 0
            regional_list.append({
                "county": node_name,
                "count": regional_data[node_name]["count"],
                "average": avg,
                "max": max_val,
                "min": min_val
            })
        
        # 按平均值排序
        regional_list.sort(key=lambda x: x["average"], reverse=True)
        return regional_list
    
    def get_weather_relationship(self):
        """获取气象数据与数量的关系"""
        result = self.read_raw_data(limit=10000)
        if result["status"] != "success":
            return []
        
        weather_cols = ["AT", "MaxT", "MinT", "AWS", "MSWS", "ADP", "Precip"]
        node_cols = [col for col in result["headers"] if col.startswith("Node_")]
        
        relationships = []
        
        for weather_col in weather_cols:
            if weather_col not in result["headers"]:
                continue
            
            for node_col in node_cols:
                weather_values = []
                node_values = []
                
                for row in result["data"]:
                    weather_val = row.get(weather_col)
                    node_val = row.get(node_col)
                    
                    if weather_val is not None and node_val is not None:
                        try:
                            weather_values.append(float(weather_val))
                            node_values.append(float(node_val))
                        except:
                            pass
                
                if len(weather_values) > 0:
                    # 计算相关系数（简单的协方差计算）
                    avg_weather = sum(weather_values) / len(weather_values)
                    avg_node = sum(node_values) / len(node_values)
                    covariance = sum([(weather_values[i] - avg_weather) * (node_values[i] - avg_node) 
                                      for i in range(len(weather_values))]) / len(weather_values) if len(weather_values) > 0 else 0
                    
                    relationships.append({
                        "weather": weather_col,
                        "region": node_col.replace("Node_", ""),
                        "correlation": covariance,
                        "avg_weather": avg_weather,
                        "avg_region": avg_node
                    })
        
        return relationships
    
    def read_prediction_data(self, model_name):
        """读取预测数据"""
        try:
            import openpyxl
            
            if model_name not in self.model_files:
                return {"status": "error", "message": f"模型 {model_name} 不存在"}
            
            file_path = os.path.join(self.prediction_data_dir, self.model_files[model_name])
            if not os.path.exists(file_path):
                return {"status": "error", "message": f"文件不存在: {file_path}"}
            
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # 读取表头
            headers = []
            for cell in ws[1]:
                headers.append(cell.value if cell.value else "")
            
            # 读取数据行
            data = []
            for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 1001)):
                row_data = {}
                for idx, cell in enumerate(row):
                    if idx < len(headers):
                        value = cell.value
                        if isinstance(value, datetime):
                            value = value.strftime('%Y-%m-%d')
                        row_data[headers[idx]] = value
                data.append(row_data)
            
            return {
                "status": "success",
                "headers": headers,
                "data": data,
                "total_rows": ws.max_row - 1
            }
        except Exception as e:
            return {
                "status": "error",
                "message": str(e),
                "headers": [],
                "data": [],
                "total_rows": 0
            }
    
    def list_prediction_models(self):
        """列出所有预测模型"""
        available_models = []
        for model_name, file_name in self.model_files.items():
            file_path = os.path.join(self.prediction_data_dir, file_name)
            if os.path.exists(file_path):
                available_models.append(model_name)
        
        return available_models
    
    def get_model_prediction_stats(self, model_name):
        """获取模型预测统计数据"""
        result = self.read_prediction_data(model_name)
        if result["status"] != "success":
            return {
                "status": "error",
                "message": result.get("message", "读取失败")
            }
        
        data = result["data"]
        if not data:
            return {
                "status": "success",
                "model": model_name,
                "total_predictions": 0,
                "stats": {}
            }
        
        # 统计预测数据
        stats = {
            "total_predictions": len(data),
            "avg_prediction": 0,
            "max_prediction": 0,
            "min_prediction": 0,
            "columns": []
        }
        
        # 找到数值列
        if data and len(data) > 0:
            numeric_cols = []
            for col in result["headers"]:
                values = [d.get(col, 0) for d in data if d.get(col)]
                if values:
                    try:
                        numeric_vals = [float(v) for v in values if v is not None]
                        if numeric_vals:
                            numeric_cols.append({
                                "name": col,
                                "values": numeric_vals,
                                "avg": sum(numeric_vals) / len(numeric_vals),
                                "max": max(numeric_vals),
                                "min": min(numeric_vals)
                            })
                    except:
                        pass
            
            if numeric_cols:
                stats["columns"] = numeric_cols
                all_values = [v for col in numeric_cols for v in col["values"]]
                stats["avg_prediction"] = sum(all_values) / len(all_values) if all_values else 0
                stats["max_prediction"] = max(all_values) if all_values else 0
                stats["min_prediction"] = min(all_values) if all_values else 0
        
        return {
            "status": "success",
            "model": model_name,
            "total_predictions": len(data),
            "stats": stats
        }
    
    def compare_models(self, model_names=None):
        """对比多个模型的预测结果"""
        if model_names is None:
            model_names = self.list_prediction_models()
        
        comparison = {
            "status": "success",
            "models": [],
            "summary": {}
        }
        
        stats_list = []
        for model_name in model_names:
            stats = self.get_model_prediction_stats(model_name)
            if stats["status"] == "success":
                stats_list.append({
                    "name": model_name,
                    "count": stats["total_predictions"],
                    "avg": stats["stats"].get("avg_prediction", 0),
                    "max": stats["stats"].get("max_prediction", 0),
                    "min": stats["stats"].get("min_prediction", 0)
                })
        
        comparison["models"] = stats_list
        
        # 生成汇总统计
        if stats_list:
            total_count = sum(s["count"] for s in stats_list)
            avg_prediction = sum(s["avg"] for s in stats_list) / len(stats_list) if stats_list else 0
            comparison["summary"] = {
                "total_models": len(stats_list),
                "total_predictions": total_count,
                "avg_prediction": avg_prediction
            }
        
        return comparison
    
    def compare_all_models(self):
        """对比所有模型（别名方法，用于兼容）"""
        return self.compare_models()
    
    def get_district_model_comparison(self):
        """获取按区县分组的模型对比数据（含真实数据）- 宽格式"""
        try:
            import openpyxl
            from collections import defaultdict
            
            # 读取原始数据
            wb = openpyxl.load_workbook(self.raw_data_path, data_only=True)
            ws = wb.active
            
            # 读取表头
            headers = [cell.value for cell in ws[1]]
            
            # 找到Date列
            date_idx = None
            for i, h in enumerate(headers):
                if h and ('Date' in str(h) or 'date' in str(h) or '日期' in str(h)):
                    date_idx = i
                    break
            
            if date_idx is None:
                return {'status': 'error', 'message': f'未找到日期列，表头: {headers[:5]}...'}
            
            # 找到所有Node_XXX列（区县列）
            district_columns = {}
            for i, h in enumerate(headers):
                if h and 'Node_' in str(h):
                    # 将Node_DaXing转换为大兴区
                    district_name = str(h).replace('Node_', '')
                    district_map = {
                        'DaXing': '大兴区',
                        'MiYun': '密云区',
                        'PingGu': '平谷区',
                        'YanQing': '延庆区',
                        'HuaiRou': '怀柔区',
                        'FangShan': '房山区',
                        'ChangPing': '昌平区',
                        'HaiDian': '海淀区',
                        'TongZhou': '通州区',
                        'ShunYi': '顺义区'
                    }
                    cn_name = district_map.get(district_name, district_name)
                    district_columns[cn_name] = i
            
            if not district_columns:
                return {'status': 'error', 'message': f'未找到区县列（Node_XXX），表头: {headers}'}
            
            # 按区县组织真实数据（宽格式转长格式）
            real_data_by_district = defaultdict(lambda: {'dates': [], 'values': []})
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                date_val = row[date_idx]
                
                if date_val:
                    # 处理日期格式
                    if isinstance(date_val, datetime):
                        date_str = date_val.strftime('%Y-%m-%d')
                    else:
                        date_str = str(date_val).split()[0]  # 去除时间部分
                    
                    # 为每个区县添加这一天的数据
                    for district_name, col_idx in district_columns.items():
                        value_val = row[col_idx]
                        if value_val is not None:
                            real_data_by_district[district_name]['dates'].append(date_str)
                            real_data_by_district[district_name]['values'].append(float(value_val) if value_val else 0)
            
            # 读取所有模型的预测数据（也是宽格式）
            model_names = self.list_prediction_models()
            model_data_by_district = defaultdict(lambda: defaultdict(lambda: {'dates': [], 'values': []}))
            
            for model_name in model_names:
                try:
                    # 使用model_files映射获取正确的文件名
                    file_name = self.model_files.get(model_name)
                    if not file_name:
                        print(f"模型 {model_name} 没有文件名映射")
                        continue
                    
                    pred_file = os.path.join(self.prediction_data_dir, file_name)
                    if not os.path.exists(pred_file):
                        print(f"模型文件不存在: {pred_file}")
                        continue
                    
                    wb_pred = openpyxl.load_workbook(pred_file, data_only=True)
                    ws_pred = wb_pred.active
                    
                    # 读取预测数据表头
                    pred_headers = [cell.value for cell in ws_pred[1]]
                    
                    # 找到Date列
                    pred_date_idx = None
                    for i, h in enumerate(pred_headers):
                        if h and ('Date' in str(h) or 'date' in str(h) or '日期' in str(h)):
                            pred_date_idx = i
                            break
                    
                    if pred_date_idx is None:
                        print(f"模型 {model_name} 缺少日期列")
                        wb_pred.close()
                        continue
                    
                    # 找到所有Node_XXX列
                    pred_district_columns = {}
                    for i, h in enumerate(pred_headers):
                        if h and 'Node_' in str(h):
                            district_name = str(h).replace('Node_', '')
                            district_map = {
                                'DaXing': '大兴区',
                                'MiYun': '密云区',
                                'PingGu': '平谷区',
                                'YanQing': '延庆区',
                                'HuaiRou': '怀柔区',
                                'FangShan': '房山区',
                                'ChangPing': '昌平区',
                                'HaiDian': '海淀区',
                                'TongZhou': '通州区',
                                'ShunYi': '顺义区'
                            }
                            cn_name = district_map.get(district_name, district_name)
                            pred_district_columns[cn_name] = i
                    
                    if not pred_district_columns:
                        print(f"模型 {model_name} 缺少区县列")
                        wb_pred.close()
                        continue
                    
                    # 读取预测数据（宽格式转长格式）
                    for row in ws_pred.iter_rows(min_row=2, values_only=True):
                        pred_date = row[pred_date_idx]
                        
                        if pred_date:
                            if isinstance(pred_date, datetime):
                                pred_date_str = pred_date.strftime('%Y-%m-%d')
                            else:
                                pred_date_str = str(pred_date).split()[0]
                            
                            # 为每个区县添加预测数据
                            for district_name, col_idx in pred_district_columns.items():
                                pred_value = row[col_idx]
                                if pred_value is not None:
                                    model_data_by_district[district_name][model_name]['dates'].append(pred_date_str)
                                    model_data_by_district[district_name][model_name]['values'].append(float(pred_value) if pred_value else 0)
                    
                    wb_pred.close()
                except Exception as e:
                    print(f"读取模型 {model_name} 数据失败: {e}")
                    import traceback
                    traceback.print_exc()
                    continue
            
            wb.close()
            
            # 组织最终数据结构
            districts = list(set(list(real_data_by_district.keys()) + list(model_data_by_district.keys())))
            districts.sort()
            
            district_data = {}
            for district in districts:
                district_data[district] = {
                    'real_data': real_data_by_district.get(district, {'dates': [], 'values': []}),
                    'models': dict(model_data_by_district.get(district, {}))
                }
            
            return {
                'status': 'success',
                'data': {
                    'districts': districts,
                    'district_data': district_data
                }
            }
        
        except Exception as e:
            import traceback
            traceback.print_exc()
            return {'status': 'error', 'message': str(e)}


if __name__ == '__main__':
    reader = SimpleDataReader()
    print("读取原始数据...")
    result = reader.read_raw_data(limit=10)
    print(f"状态: {result['status']}")
    print(f"表头: {result['headers']}")
    print(f"数据行数: {len(result['data'])}")
    if result['data']:
        print("第一条数据:", result['data'][0])
    
    print("\n年度统计:")
    yearly = reader.get_yearly_statistics()
    for item in yearly[:5]:
        print(item)
    
    print("\n月度统计:")
    monthly = reader.get_monthly_statistics()
    for item in monthly[:5]:
        print(item)
    
    print("\n区域统计:")
    regional = reader.get_regional_statistics()
    for item in regional[:5]:
        print(item)
    
    print("\n可用预测模型:")
    models = reader.list_prediction_models()
    print(models)

